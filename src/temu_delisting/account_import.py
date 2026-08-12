"""从一份 Excel 清单批量导入账号"架子"。

只导入名称类信息（不可能导入登录态/Cookie——那是登录之后才有的会话数据，
没法提前写在一份表格里），导入完的账号条目登录状态都是空的，还需要对每个
条目单独走"+ 添加账号"（全新登录）或"+ 复制账号"（跟别的条目共用登录）
去补登录信息。

期望的表格列（按表头文字找列，不认死列的顺序）：
    账号        —— 手机号，纯展示/分组用，不是登录凭证，不会拿去做任何
                    自动化操作。绝对不要放 Temu 登录密码这类真实凭证进
                    表格——系统本身做不到自动登录，密码导进来也用不上，
                    只会变成明文躺在共享表格和账号数据库里，是真实的
                    泄露风险，登录必须由知道密码的人手动操作。
    店铺数量     —— 这个手机号下一共几个店铺
    全托管       —— 全托管店铺数（这张表只给了"手机号级别"的汇总数字，
                    没法知道具体是底下哪几个店铺，所以只记进备注，不会
                    分摊到某个具体店铺上）
    半托管       —— 同上
    店铺编号     —— 内部编号区间（比如"店铺01-01~02"），纯展示用，不是
                    网页上真实显示的店铺名字，不会当成 mall_name 用
    店铺名称     —— 网页上真实显示的店铺名字，用顿号/逗号/斜杠隔开，
                    拆开之后每一个就是一个要创建的账号条目（display_name
                    和 mall_name 都用这个，因为 mall_name 必须跟网页上
                    显示的文字完全一致才能用来自动切换店铺）
    备注        —— 可选，自己想写什么都行（谁负责的、什么时候接手的之类），
                    原样保留在导入后账号的"备注"字段最前面，不会被系统
                    自动生成的那部分覆盖掉
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

_HEADER_ALIASES = {
    "账号": "phone",
    "店铺数量": "shop_count",
    "全托管": "full_managed",
    "半托管": "semi_managed",
    "店铺编号": "shop_code",
    "店铺名称": "shop_names",
    "备注": "custom_notes",
}
_REQUIRED_KEYS = {"phone", "shop_count", "shop_names"}

_NAME_SPLIT_PATTERN = re.compile(r"[、,，/]+")

IMPORT_FORMAT_HELP = """Excel 表格第一行是表头，列名要写得跟下面完全一致（顺序无所谓，多余的列会被忽略）：

必须要有的列：
・账号 —— 手机号，纯展示/分组用，不是登录密码，绝对不要把 Temu 登录密码写进这张表
・店铺数量 —— 这个手机号下一共几个店铺
・店铺名称 —— 网页上真实显示的店铺名字，多个店铺用顿号、逗号或斜杠隔开，比如"SaveNest、Dwmane Shop"

可以不填的列：
・全托管 / 半托管 —— 数量，会自动记进备注里
・店铺编号 —— 内部编号，纯展示用
・备注 —— 自己想写什么都行，会原样保留在导入后账号的备注最前面

导入进来的账号只有"架子"（名字、备注），还没有登录信息，每个账号导入后
还要单独用「+ 添加账号」或「+ 复制账号」手动登录一次。"""


@dataclass
class ImportedShop:
    display_name: str
    mall_name: str
    notes: str


@dataclass
class ImportRowIssue:
    row_number: int
    message: str


@dataclass
class ImportResult:
    shops: list[ImportedShop] = field(default_factory=list)
    issues: list[ImportRowIssue] = field(default_factory=list)


def parse_account_excel(path: Path) -> ImportResult:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ImportResult()

    column_index: dict[str, int] = {}
    for i, cell in enumerate(rows[0]):
        key = _HEADER_ALIASES.get(str(cell).strip() if cell is not None else "")
        if key:
            column_index[key] = i

    missing = _REQUIRED_KEYS - column_index.keys()
    if missing:
        raise ValueError(
            f"表头里缺少必须的列：{'、'.join(missing)}——至少要有「账号」「店铺数量」"
            "「店铺名称」这三列，列名要完全一致（顺序不影响）"
        )

    result = ImportResult()
    for row_number, row in enumerate(rows[1:], start=2):
        def cell(key: str):
            idx = column_index.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        phone_raw = cell("phone")
        shop_names_raw = cell("shop_names")
        if phone_raw is None and shop_names_raw is None:
            continue  # 空行

        phone_text = str(phone_raw).strip() if phone_raw is not None else ""

        shop_count_raw = cell("shop_count")
        try:
            shop_count = int(shop_count_raw) if shop_count_raw is not None else None
        except (TypeError, ValueError):
            shop_count = None

        full_managed = cell("full_managed")
        semi_managed = cell("semi_managed")
        custom_notes_raw = cell("custom_notes")
        custom_notes = str(custom_notes_raw).strip() if custom_notes_raw is not None else ""

        names = [n.strip() for n in _NAME_SPLIT_PATTERN.split(str(shop_names_raw or "").strip()) if n.strip()]
        if not names:
            result.issues.append(
                ImportRowIssue(row_number, f"手机号 {phone_text} 这一行没解析到店铺名称，已跳过")
            )
            continue

        if shop_count is not None and len(names) != shop_count:
            result.issues.append(
                ImportRowIssue(
                    row_number,
                    f"手机号 {phone_text}：店铺数量写的是 {shop_count}，但店铺名称实际解析出 "
                    f"{len(names)} 个，已按实际解析出的 {len(names)} 个导入，请核对源文件",
                )
            )

        note_parts = []
        if phone_text:
            note_parts.append(f"导入自手机号 {phone_text}")
        if shop_count is not None:
            note_parts.append(f"该手机号下共 {shop_count} 个店铺")
        if full_managed is not None or semi_managed is not None:
            note_parts.append(
                f"全托管 {full_managed if full_managed is not None else '?'}/"
                f"半托管 {semi_managed if semi_managed is not None else '?'}"
            )
        auto_notes = "，".join(note_parts)
        # 自己写的备注放最前面，不会被自动生成的那部分挤掉或覆盖。
        notes = f"{custom_notes}（{auto_notes}）" if custom_notes and auto_notes else (custom_notes or auto_notes)

        for name in names:
            result.shops.append(ImportedShop(display_name=name, mall_name=name, notes=notes))

    return result
